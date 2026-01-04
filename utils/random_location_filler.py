"""Fill missing locations in parsed_records.xlsx using LOCATIONS from .env."""
from __future__ import annotations

import argparse
import random
from pathlib import Path
from typing import Optional

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover - optional dependency
    raise SystemExit(
        "openpyxl is required to edit parsed_records.xlsx in place. Install via `pip install openpyxl`."
    ) from exc

from configs.config import FOLDER_PATH, LOCATIONS


def _default_input() -> Path:
    return FOLDER_PATH / "parsed_records.xlsx"


def _needs_location(value: object) -> bool:
    if value is None:
        return True
    text = str(value).strip()
    if not text:
        return True
    return text.lower() in {"nan", "none", "null"}


def _find_location_column(header_cells: list[object]) -> Optional[int]:
    for idx, value in enumerate(header_cells, start=1):
        if value is None:
            continue
        name = str(value).strip().lower()
        if name == "location":
            return idx
    return None


def fill_locations(input_path: Path, output_path: Path, seed: Optional[int]) -> tuple[int, int]:
    if not LOCATIONS:
        raise ValueError("LOCATIONS is empty. Check LOCATION_GROUPS in .env.")
    if not input_path.exists():
        raise FileNotFoundError(f"parsed_records.xlsx not found: {input_path}")

    rng = random.Random(seed)
    workbook = load_workbook(input_path)
    sheet = workbook.active

    header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    location_col = _find_location_column(header)
    if location_col is None:
        location_col = len(header) + 1
        sheet.cell(row=1, column=location_col, value="location")

    filled = 0
    total = 0
    for row in range(2, sheet.max_row + 1):
        total += 1
        cell = sheet.cell(row=row, column=location_col)
        if _needs_location(cell.value):
            cell.value = rng.choice(LOCATIONS)
            filled += 1

    workbook.save(output_path)
    return filled, total


def main() -> int:
    parser = argparse.ArgumentParser(description="Fill missing locations in parsed_records.xlsx.")
    parser.add_argument("--input", default="", help="Input parsed_records.xlsx path.")
    parser.add_argument("--output", default="", help="Output path (default: overwrite input).")
    parser.add_argument("--seed", type=int, default=0, help="Random seed (0 = random).")
    args = parser.parse_args()

    input_path = Path(args.input).expanduser().resolve() if args.input else _default_input()
    output_path = Path(args.output).expanduser().resolve() if args.output else input_path
    seed = None if args.seed == 0 else int(args.seed)

    try:
        filled, total = fill_locations(input_path, output_path, seed)
    except (ValueError, FileNotFoundError) as exc:
        print(str(exc))
        return 2

    print(f"Updated {filled} of {total} rows.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
